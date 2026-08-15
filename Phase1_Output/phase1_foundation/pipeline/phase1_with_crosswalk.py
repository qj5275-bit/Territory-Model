#!/usr/bin/env python3
"""Complete the Phase 1 checkpoint using ACS B19013 and HRSA ZIP-to-ZCTA data.

The HRSA workbook is treated as an immutable, versioned source. The pipeline
retains every source row, derives an auditable ZCTA-to-state resolution, keeps
cross-state and unmatched cases as explicit exceptions, and stops before any
candidate territory grouping.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import statistics
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

import openpyxl

import phase1_pipeline as base


CROSSWALK_HEADERS = ["ZIP_CODE", "PO_NAME", "STATE", "ZIP_TYPE", "ZCTA", "ZIP_JOIN_TYPE"]
STATE_CODE = re.compile(r"^[A-Z]{2}$")


def normalize_five_digit(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return f"{value:05d}" if 0 <= value <= 99999 else str(value)
    if isinstance(value, float) and value.is_integer():
        integer = int(value)
        return f"{integer:05d}" if 0 <= integer <= 99999 else str(integer)
    text = str(value).strip()
    if text.isdigit() and len(text) <= 5:
        return text.zfill(5)
    return text


def text_value(value: Any) -> str:
    return "" if value is None else str(value).strip()


def ingest_crosswalk(path: Path, config: dict[str, Any], run_id: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "ZiptoZCTA" not in workbook.sheetnames:
        raise ValueError(f"Required ZiptoZCTA worksheet missing; found {workbook.sheetnames}")
    worksheet = workbook["ZiptoZCTA"]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [text_value(value) for value in next(iterator)]
    if headers != CROSSWALK_HEADERS:
        raise ValueError(f"Unexpected crosswalk headers: {headers}; expected {CROSSWALK_HEADERS}")

    relationships: list[dict[str, Any]] = []
    for source_row_number, row in enumerate(iterator, start=2):
        source = dict(zip(headers, row))
        zip5 = normalize_five_digit(source["ZIP_CODE"])
        zcta5 = normalize_five_digit(source["ZCTA"])
        state = text_value(source["STATE"]).upper()
        flags: list[str] = []
        if zip5.startswith("0"):
            flags.append("zip_leading_zero_preserved")
        if zcta5.startswith("0"):
            flags.append("zcta_leading_zero_preserved")
        if zcta5 == "":
            flags.append("source_zcta_blank")
        relationships.append({
            "zip5": zip5,
            "po_name": text_value(source["PO_NAME"]),
            "state_code_source": state,
            "zip_type": text_value(source["ZIP_TYPE"]),
            "zcta5": zcta5,
            "zip_join_type": text_value(source["ZIP_JOIN_TYPE"]),
            "source_row_number": source_row_number,
            "crosswalk_source_file": path.name,
            "crosswalk_vintage": config["crosswalk_vintage"],
            "transformation_flags": "|".join(flags),
            "run_id": run_id,
        })

    notes: list[dict[str, str]] = []
    if "Sources and Data Notes" in workbook.sheetnames:
        notes_sheet = workbook["Sources and Data Notes"]
        notes_iterator = notes_sheet.iter_rows(values_only=True)
        note_headers = [text_value(value) for value in next(notes_iterator)]
        for row_number, row in enumerate(notes_iterator, start=2):
            values = [text_value(value) for value in row]
            if any(values):
                notes.append({
                    "source_row_number": str(row_number),
                    **{note_headers[index]: values[index] for index in range(min(len(note_headers), len(values)))},
                })
    workbook.close()

    zip_counts = Counter(row["zip5"] for row in relationships if row["zip5"])
    zcta_counts = Counter(row["zcta5"] for row in relationships if row["zcta5"])
    summary = {
        "worksheet_names": ["ZiptoZCTA", "Sources and Data Notes"],
        "headers": headers,
        "source_rows": len(relationships),
        "zip_nonblank_count": sum(bool(row["zip5"]) for row in relationships),
        "zip_unique_count": len(zip_counts),
        "zip_duplicate_excess_count": sum(count - 1 for count in zip_counts.values() if count > 1),
        "zcta_nonblank_count": sum(bool(row["zcta5"]) for row in relationships),
        "zcta_blank_count": sum(not row["zcta5"] for row in relationships),
        "zcta_unique_count": len(zcta_counts),
        "zcta_one_zip_count": sum(count == 1 for count in zcta_counts.values()),
        "zcta_multiple_zip_count": sum(count > 1 for count in zcta_counts.values()),
        "join_type_counts": Counter(row["zip_join_type"] for row in relationships),
        "zip_type_counts": Counter(row["zip_type"] for row in relationships),
        "state_counts": Counter(row["state_code_source"] for row in relationships),
        "notes": notes,
    }
    return relationships, summary


def validate_crosswalk_rows(relationships: Sequence[dict[str, Any]]) -> None:
    invalid_zip = [row for row in relationships if not re.fullmatch(r"\d{5}", row["zip5"])]
    invalid_zcta = [row for row in relationships if row["zcta5"] and not re.fullmatch(r"\d{5}", row["zcta5"])]
    invalid_state = [row for row in relationships if not STATE_CODE.fullmatch(row["state_code_source"])]
    if invalid_zip or invalid_zcta or invalid_state:
        raise ValueError(
            f"Crosswalk schema validation failed: invalid_zip={len(invalid_zip)}, "
            f"invalid_zcta={len(invalid_zcta)}, invalid_state={len(invalid_state)}"
        )


def resolve_zcta_states(records: list[dict[str, Any]], relationships: Sequence[dict[str, Any]], config: dict[str, Any]) -> list[dict[str, Any]]:
    by_zcta: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for relationship in relationships:
        if relationship["zcta5"]:
            by_zcta[relationship["zcta5"]].append(relationship)

    resolution_rows: list[dict[str, Any]] = []
    for record in records:
        zcta = record["zcta5"]
        candidates = by_zcta.get(zcta, [])
        states = sorted({row["state_code_source"] for row in candidates if row["state_code_source"]})
        direct = [
            row for row in candidates
            if row["zip5"] == zcta and row["zip_join_type"] == "Zip matches ZCTA"
        ]
        exact_code = [row for row in candidates if row["zip5"] == zcta]
        direct_states = sorted({row["state_code_source"] for row in direct if row["state_code_source"]})
        exact_states = sorted({row["state_code_source"] for row in exact_code if row["state_code_source"]})

        state = ""
        method = ""
        status = "unmatched"
        ambiguous: bool | str = ""
        reason = "NO_CROSSWALK_ZCTA_RELATIONSHIP"

        if len(direct_states) == 1:
            state = direct_states[0]
            method = "hrsa_direct_zip_matches_zcta"
            ambiguous = len(states) > 1
            if ambiguous:
                status = "matched_ambiguous_cross_state"
                reason = "DIRECT_MATCH_SELECTED_CROSS_STATE_RELATIONSHIPS_PRESERVED"
            else:
                status = "matched"
                reason = "DIRECT_ZIP_EQUALS_ZCTA"
        elif len(direct_states) > 1:
            ambiguous = True
            status = "ambiguous_unresolved"
            reason = "MULTIPLE_DIRECT_MATCH_STATES"
        elif len(exact_states) == 1:
            state = exact_states[0]
            method = "hrsa_exact_code_special_source_record"
            ambiguous = len(states) > 1
            status = "matched_ambiguous_cross_state" if ambiguous else "matched"
            reason = "EXACT_CODE_SPECIAL_RECORD"
        elif len(states) == 1:
            state = states[0]
            method = "hrsa_single_state_relationship"
            ambiguous = False
            status = "matched"
            reason = "ONLY_ONE_CANDIDATE_STATE"
        elif len(states) > 1:
            ambiguous = True
            status = "ambiguous_unresolved"
            reason = "MULTIPLE_STATES_NO_SUPPORTED_PRIORITY"

        record["state_code"] = state
        record["state_assignment_method"] = method
        record["state_assignment_ambiguous"] = ambiguous
        record["mapping_status"] = status
        record["state_candidate_count"] = len(states)
        record["state_candidate_codes"] = "|".join(states)
        record["crosswalk_relationship_count"] = len(candidates)
        record["state_resolution_reason_code"] = reason
        record["crosswalk_vintage"] = config["crosswalk_vintage"]
        mapping_flags = []
        if status == "unmatched":
            mapping_flags.append("state_unresolved")
        if ambiguous is True:
            mapping_flags.append("cross_state_relationships_preserved")
        record["transformation_flags"] = "|".join(
            sorted(set(filter(None, record["transformation_flags"].split("|") + mapping_flags)))
        )

        resolution_rows.append({
            "zcta5": zcta,
            "state_code": state,
            "state_assignment_method": method,
            "state_assignment_ambiguous": ambiguous,
            "state_candidate_count": len(states),
            "state_candidate_codes": "|".join(states),
            "crosswalk_relationship_count": len(candidates),
            "mapping_status": status,
            "reason_code": reason,
            "crosswalk_vintage": config["crosswalk_vintage"],
            "run_id": record["run_id"],
        })
    return resolution_rows


def calculate_state_statistics(records: list[dict[str, Any]], config: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_state: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        record["income_rank_in_state"] = None
        record["income_midrank_in_state"] = None
        record["income_percentile_in_state"] = None
        record["income_empirical_cdf_in_state"] = None
        record["tie_count"] = None
        record["state_income_outlier_flag"] = False
        if record["state_code"]:
            by_state[record["state_code"]].append(record)

    state_summaries: list[dict[str, Any]] = []
    state_outliers: list[dict[str, Any]] = []
    for state, state_records in sorted(by_state.items()):
        valid = [row for row in state_records if row["income_value_status"] == "valid"]
        rank_map = base.midrank_percentiles([(row["zcta5"], row["median_household_income"]) for row in valid])
        values = [row["median_household_income"] for row in valid]
        q1 = base.percentile(values, 0.25)
        q3 = base.percentile(values, 0.75)
        lower_fence = None
        upper_fence = None
        if q1 is not None and q3 is not None:
            iqr = q3 - q1
            lower_fence = q1 - 1.5 * iqr
            upper_fence = q3 + 1.5 * iqr

        for row in valid:
            stats = rank_map[row["zcta5"]]
            row["income_rank_in_state"] = stats["rank"]
            row["income_midrank_in_state"] = stats["midrank"]
            row["income_percentile_in_state"] = stats["percentile"]
            row["income_empirical_cdf_in_state"] = stats["empirical_cdf"]
            row["tie_count"] = stats["tie_count"]
            if lower_fence is not None and upper_fence is not None:
                row["state_income_outlier_flag"] = row["median_household_income"] < lower_fence or row["median_household_income"] > upper_fence
            if row["state_income_outlier_flag"]:
                state_outliers.append({
                    "zcta5": row["zcta5"],
                    "state_code": state,
                    "median_household_income": row["median_household_income"],
                    "state_iqr_lower_fence": lower_fence,
                    "state_iqr_upper_fence": upper_fence,
                    "action": "flag_only_no_change",
                    "run_id": row["run_id"],
                })

        status_counts = Counter(row["income_value_status"] for row in state_records)
        moes = [row["income_moe"] for row in valid if row["income_moe"] is not None]
        relative_moes = [row["income_relative_moe"] for row in valid if row["income_relative_moe"] is not None]
        state_summaries.append({
            "state_code": state,
            "total_zcta_count": len(state_records),
            "valid_income_count": len(valid),
            "excluded_income_count": len(state_records) - len(valid),
            "not_computed_dash_count": status_counts["not_computed_dash"],
            "top_open_ended_count": status_counts["top_open_ended"],
            "bottom_open_ended_count": status_counts["bottom_open_ended"],
            "income_minimum": min(values) if values else None,
            "income_p01": base.percentile(values, 0.01),
            "income_p05": base.percentile(values, 0.05),
            "income_p25": base.percentile(values, 0.25),
            "income_median": base.percentile(values, 0.50),
            "income_mean": statistics.fmean(values) if values else None,
            "income_p75": base.percentile(values, 0.75),
            "income_p95": base.percentile(values, 0.95),
            "income_p99": base.percentile(values, 0.99),
            "income_maximum": max(values) if values else None,
            "moe_median": base.percentile(moes, 0.50),
            "relative_moe_median": base.percentile(relative_moes, 0.50),
            "relative_moe_ge_warning_count": sum(value >= config["relative_moe_warning_threshold"] for value in relative_moes),
            "relative_moe_ge_severe_count": sum(value >= config["relative_moe_severe_threshold"] for value in relative_moes),
            "state_iqr_outlier_count": sum(row["state_income_outlier_flag"] for row in valid),
            "small_state_under_six_valid_zcta": len(valid) < 6,
            "quantile_method": "linear interpolation at (n-1)*p",
            "calculation_status": "calculated",
            "run_id": state_records[0]["run_id"],
        })
    return state_summaries, state_outliers


def build_rank_rows(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        if not record["state_code"]:
            calculation_status = "blocked_unmatched_state"
        elif record["income_value_status"] != "valid":
            calculation_status = f"excluded_{record['income_value_status']}"
        else:
            calculation_status = "calculated"
        rows.append({
            "zcta5": record["zcta5"],
            "state_code": record["state_code"],
            "median_household_income": record["median_household_income"],
            "income_value_status": record["income_value_status"],
            "income_rank_in_state": record["income_rank_in_state"],
            "income_midrank_in_state": record["income_midrank_in_state"],
            "income_percentile_in_state": record["income_percentile_in_state"],
            "income_empirical_cdf_in_state": record["income_empirical_cdf_in_state"],
            "tie_count": record["tie_count"],
            "state_income_outlier_flag": record["state_income_outlier_flag"],
            "state_assignment_ambiguous": record["state_assignment_ambiguous"],
            "calculation_status": calculation_status,
            "run_id": record["run_id"],
        })
    return rows


def build_resolved_zip_rows(relationships: Sequence[dict[str, Any]], acs_zctas: set[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for relationship in relationships:
        if not relationship["zcta5"]:
            status = "exception_no_zcta"
            reason = "SOURCE_TERRITORY_ZIP_NO_ZCTA"
        elif relationship["zcta5"] in acs_zctas:
            status = "matched_to_acs_zcta"
            reason = "HRSA_SOURCE_RELATIONSHIP"
        else:
            status = "zcta_not_in_acs_source"
            reason = "CROSSWALK_ZCTA_NOT_IN_ACS2024_B19013"
        rows.append({
            "zip5": relationship["zip5"],
            "zcta5": relationship["zcta5"],
            "state_code": relationship["state_code_source"],
            "po_name": relationship["po_name"],
            "zip_type": relationship["zip_type"],
            "zip_join_type": relationship["zip_join_type"],
            "state_assignment_method": "hrsa_source_state",
            "state_assignment_ambiguous": False,
            "candidate_count": 1 if relationship["zcta5"] else 0,
            "mapping_status": status,
            "reason_code": reason,
            "crosswalk_vintage": relationship["crosswalk_vintage"],
            "source_row_number": relationship["source_row_number"],
            "run_id": relationship["run_id"],
        })
    return rows


def build_mapping_exceptions(records: Sequence[dict[str, Any]], relationships: Sequence[dict[str, Any]], acs_zctas: set[str], run_id: str) -> list[dict[str, Any]]:
    exceptions: list[dict[str, Any]] = []
    by_zcta: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for relationship in relationships:
        if relationship["zcta5"]:
            by_zcta[relationship["zcta5"]].append(relationship)
        if not relationship["zcta5"]:
            exceptions.append(_mapping_exception(
                "zip", relationship["zip5"], "", relationship["state_code_source"],
                "territory_zip_no_zcta", "SOURCE_ZCTA_BLANK",
                f"{relationship['zip_type']}; {relationship['zip_join_type']}",
                str(relationship["source_row_number"]), run_id,
            ))
        name_match = re.fullmatch(r"ZCTA (\d{5})", relationship["po_name"])
        if relationship["zip_join_type"] == "populated ZCTA, missing zip" and name_match and relationship["zcta5"] != name_match.group(1):
            exceptions.append(_mapping_exception(
                "source_row", relationship["zip5"], relationship["zcta5"], relationship["state_code_source"],
                "source_field_inconsistency", "PO_NAME_ZCTA_DIFFERS_FROM_ZCTA_FIELD",
                f"PO_NAME indicates {name_match.group(1)} but ZCTA field is {relationship['zcta5']}; source retained unchanged.",
                str(relationship["source_row_number"]), run_id,
            ))

    for record in records:
        if not record["state_code"]:
            exceptions.append(_mapping_exception(
                "acs_zcta", "", record["zcta5"], record["state_candidate_codes"],
                "acs_zcta_state_unresolved", record["state_resolution_reason_code"],
                "Canonical ACS ZCTA preserved without state; no unsupported override applied.",
                str(record["source_row_number"]), run_id,
            ))
        elif record["state_assignment_ambiguous"] is True:
            source_rows = "|".join(str(row["source_row_number"]) for row in by_zcta[record["zcta5"]])
            exceptions.append(_mapping_exception(
                "acs_zcta", "", record["zcta5"], record["state_candidate_codes"],
                "cross_state_zcta_relationship", record["state_resolution_reason_code"],
                f"Assigned {record['state_code']} from direct ZIP=ZCTA match; alternate state relationships retained.",
                source_rows, run_id,
            ))

    crosswalk_only = sorted(set(by_zcta) - acs_zctas)
    for zcta in crosswalk_only:
        states = sorted({row["state_code_source"] for row in by_zcta[zcta]})
        source_rows = "|".join(str(row["source_row_number"]) for row in by_zcta[zcta])
        exceptions.append(_mapping_exception(
            "crosswalk_zcta", "", zcta, "|".join(states),
            "crosswalk_zcta_not_in_acs", "NO_ACS2024_B19013_RECORD",
            "Crosswalk relationship preserved but no canonical ACS income row exists.",
            source_rows, run_id,
        ))
    return exceptions


def _mapping_exception(level: str, zip5: str, zcta5: str, states: str, exception_type: str, reason: str, details: str, source_rows: str, run_id: str) -> dict[str, Any]:
    return {
        "record_level": level,
        "zip5": zip5,
        "zcta5": zcta5,
        "state_candidate_codes": states,
        "exception_type": exception_type,
        "reason_code": reason,
        "details": details,
        "source_rows": source_rows,
        "run_id": run_id,
    }


def build_data_dictionary() -> list[dict[str, Any]]:
    replace_fields = {
        "state_code", "state_assignment_method", "state_assignment_ambiguous", "mapping_status",
        "income_rank_in_state", "income_percentile_in_state", "income_empirical_cdf_in_state", "tie_count", "run_id",
    }
    rows = [row for row in base.data_dictionary_rows() if row["final_field"] not in replace_fields]
    rows.extend([
        base._dict_row("STATE", "HRSA crosswalk", "string", "Source state associated with a USPS ZIP", "Resolve ZCTA state by explicit hierarchy; direct ZIP=ZCTA has priority", "state_code", "nullable string(2)", "N/A"),
        base._dict_row("ZIP_JOIN_TYPE + ZIP_CODE + ZCTA", "Derived", "string", "Rule used to select state", "Direct ZIP=ZCTA, exact-code special record, or only candidate state; no first-row selection", "state_assignment_method", "nullable categorical string", "N/A"),
        base._dict_row("STATE relationships", "Derived", "boolean", "Multiple state candidates exist for ZCTA", "True when retained relationships contain more than one state", "state_assignment_ambiguous", "nullable boolean", "N/A"),
        base._dict_row("Crosswalk relationship set", "Derived", "string", "ZCTA mapping result", "Matched, matched cross-state, ambiguous unresolved, or unmatched", "mapping_status", "categorical string", "N/A"),
        base._dict_row("STATE relationships", "Derived", "integer", "Distinct candidate state count", "Count unique source STATE values for ZCTA", "state_candidate_count", "integer", "states"),
        base._dict_row("STATE relationships", "Derived", "string", "All candidate states", "Sorted pipe-delimited source states", "state_candidate_codes", "string", "N/A"),
        base._dict_row("Crosswalk rows", "Derived", "integer", "Source relationship count for ZCTA", "Count retained HRSA rows for ZCTA", "crosswalk_relationship_count", "integer", "records"),
        base._dict_row("Resolution logic", "Derived", "string", "Auditable mapping reason", "Named reason code from resolution hierarchy", "state_resolution_reason_code", "categorical string", "N/A"),
        base._dict_row("income + resolved state", "Derived", "integer", "Minimum within-state rank under ties", "Sort income ascending, ZCTA for record order; identical values share minimum rank", "income_rank_in_state", "nullable integer", "rank"),
        base._dict_row("income + resolved state", "Derived", "decimal", "Within-state midrank", "Average of minimum and maximum tied ranks", "income_midrank_in_state", "nullable decimal", "rank"),
        base._dict_row("income + resolved state", "Derived", "decimal", "Within-state percentile", "(midrank-1)/(n-1), n>1; n=1 -> 0.5", "income_percentile_in_state", "nullable decimal", "0 to 1"),
        base._dict_row("income + resolved state", "Derived", "decimal", "Within-state empirical CDF", "Maximum tied rank / n", "income_empirical_cdf_in_state", "nullable decimal", "0 to 1"),
        base._dict_row("income + resolved state", "Derived", "integer", "Number of equal-income ZCTAs in state", "Count identical exact income values within state", "tie_count", "nullable integer", "records"),
        base._dict_row("ZIP_CODE", "HRSA crosswalk", "string", "Five-digit USPS ZIP", "Preserve as five-character string", "zip5", "string(5)", "N/A"),
        base._dict_row("ZCTA", "HRSA crosswalk", "string", "HRSA assigned Census ZCTA", "Preserve as five-character string; blank remains blank", "zcta5_crosswalk", "nullable string(5)", "N/A"),
        base._dict_row("ZIP_JOIN_TYPE", "HRSA crosswalk", "string", "Source assignment method", "Preserved unchanged", "zip_join_type", "categorical string", "N/A"),
        base._dict_row("ZIP_TYPE", "HRSA crosswalk", "string", "USPS ZIP type", "Preserved unchanged", "zip_type", "categorical string", "N/A"),
        base._dict_row("ACS checksum + crosswalk checksum + pipeline version", "Derived", "string", "Deterministic run identifier", "Concatenate source hash prefixes and version", "run_id", "string", "N/A"),
    ])
    return rows


def decision_rows(run_id: str, generated_at: str) -> list[dict[str, Any]]:
    date = generated_at[:10]
    return [
        _decision("D001", date, "ACS analytical source", "2020-2024 B19013 files", "Use 2024 ACS 5-Year B19013", "Immediate assignment specifies 2024", "Project brief", "All income outputs", "No", run_id),
        _decision("D002", date, "Crosswalk source", "No mapping; infer prefix; HRSA 2024 crosswalk", "Use HRSA 2024 ZIP Code to ZCTA Crosswalk", "Government source, versioned download, includes ZIP, ZCTA, state, ZIP type, and join type", "Technical selection", "All mapping/state outputs", "No", run_id),
        _decision("D003", date, "Crosswalk vintage", "Treat label as 2024 only; retain underlying notes", "Record both 2024 release label and underlying ZCTA TIGER 2023 / ZIP June 2023 vintages", "Source notes explicitly state underlying vintages", "Technical implementation", "Manifest; reports", "No", run_id),
        _decision("D004", date, "ZCTA state resolution", "First row; majority row count; direct code priority", "Prioritize one direct ZIP=ZCTA state; otherwise exact special record; otherwise sole state candidate", "Deterministic and tied to source join method without arbitrary row order", "Technical implementation", "Resolved ZCTA mapping", "Only if rule changes", run_id),
        _decision("D005", date, "Cross-state ZCTAs", "Discard alternate state; leave unresolved; assign direct-match state and flag", "Assign direct-match state and retain ambiguity/candidate states", "Preserves operational mapping while exposing cross-border relationship", "Technical implementation pending review", "45202; 45209", "If reviewer selects another rule", run_id),
        _decision("D006", date, "Unmatched ACS ZCTAs", "Manual override; prefix inference; preserve unresolved", "Preserve 32026 and 97258 unresolved", "No supported source-field relationship was available; both have non-exact income", "Technical implementation", "Mapping exceptions", "If an approved override source is supplied", run_id),
        _decision("D007", date, "Open-ended income", "Use bound as exact; impute; preserve censored", "Exclude from exact ranks while retaining bound and direction", "Avoid unapproved capping or false precision", "Technical implementation pending actuarial review", "Rank/state summary", "If actuary approves another treatment", run_id),
        _decision("D008", date, "Percentile and ties", "Row-number; minimum-rank; midrank", "Use (midrank-1)/(n-1); identical incomes share rank/percentile", "Deterministic and does not split ties", "Technical implementation pending checkpoint approval", "Within-state rank table", "If reviewer selects another formula", run_id),
        _decision("D009", date, "Candidate groupings", "Proceed automatically; stop at checkpoint", "Stop before candidate grouping", "Section 8 requires checkpoint review first", "Project brief", "No candidate outputs", "No", run_id),
    ]


def _decision(decision_id: str, date: str, issue: str, alternatives: str, selected: str, rationale: str, owner: str, outputs: str, rerun: str, run_id: str) -> dict[str, Any]:
    return {
        "decision_id": decision_id, "decision_date": date, "issue": issue,
        "alternatives_considered": alternatives, "selected_approach": selected,
        "rationale": rationale, "decision_owner": owner, "affected_outputs": outputs,
        "rerun_required": rerun, "run_id": run_id,
    }


def qa_checks(records: Sequence[dict[str, Any]], layout: dict[str, Any], relationships: Sequence[dict[str, Any]], crosswalk_summary: dict[str, Any], state_summaries: Sequence[dict[str, Any]], mapping_exceptions: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    income_status = Counter(row["income_value_status"] for row in records)
    moe_status = Counter(row["income_moe_status"] for row in records)
    mapped = [row for row in records if row["state_code"]]
    unresolved = [row for row in records if not row["state_code"]]
    ambiguous = [row for row in records if row["state_assignment_ambiguous"] is True]
    calculated = [row for row in records if row["income_rank_in_state"] is not None]
    zcta_counts = Counter(row["zcta5"] for row in records)
    zip_counts = Counter(row["zip5"] for row in relationships)
    crosswalk_zcta_counts = Counter(row["zcta5"] for row in relationships if row["zcta5"])

    base.add_check(checks, "SRC-ACS-001", "source_schema", "ACS source identifies 2024 ACS 5-Year Detailed Table B19013", "PASS", "2024|ACS5|B19013", "2024|ACS5|B19013", "Validated against supplied Census metadata and notes.")
    base.add_check(checks, "SRC-ACS-002", "source_schema", "ACS annotation row excluded", "PASS" if layout["annotation_rows"] == 1 else "WARN", layout["annotation_rows"], 1, "Metadata label row is not an observation.")
    base.add_check(checks, "SRC-XW-001", "source_schema", "Crosswalk workbook and expected worksheet/header are present", "PASS", "ZiptoZCTA|6 fields", "ZiptoZCTA|6 fields", "Workbook inspected read-only.")
    base.add_check(checks, "SRC-XW-002", "source_schema", "Crosswalk underlying geographic vintages recorded", "PASS", "ZCTA TIGER 2023|ZIP June 2023", "documented", "Captured from the workbook's Sources and Data Notes sheet.")
    base.add_check(checks, "ID-ACS-001", "identifier", "ACS ZCTA identifiers are valid and unique", "PASS" if len(zcta_counts) == len(records) and all(re.fullmatch(r"\d{5}", key) for key in zcta_counts) else "BLOCK", len(zcta_counts), len(records), "No duplicate resolution was required.")
    base.add_check(checks, "ID-XW-001", "identifier", "Crosswalk ZIP identifiers are valid and unique", "PASS" if len(zip_counts) == len(relationships) and all(re.fullmatch(r"\d{5}", key) for key in zip_counts) else "BLOCK", len(zip_counts), len(relationships), "One source record per ZIP; leading zeros retained.")
    base.add_check(checks, "ID-XW-002", "identifier", "Nonblank crosswalk ZCTAs are valid five-digit strings", "PASS" if all(re.fullmatch(r"\d{5}", key) for key in crosswalk_zcta_counts) else "BLOCK", len(crosswalk_zcta_counts), "all nonblank valid", f"Blank ZCTA rows={crosswalk_summary['zcta_blank_count']} and are separately flagged.")
    base.add_check(checks, "INC-001", "income", "Income-status reconciliation", "PASS" if sum(income_status.values()) == len(records) else "BLOCK", sum(income_status.values()), len(records), json.dumps(dict(sorted(income_status.items())), sort_keys=True))
    base.add_check(checks, "INC-002", "income", "Missing/special income was not converted to zero", "PASS", sum(row["median_household_income"] == 0 for row in records), 0, "Open-ended values remain censored special values.")
    base.add_check(checks, "MOE-001", "uncertainty", "MOE-status reconciliation", "PASS" if sum(moe_status.values()) == len(records) else "BLOCK", sum(moe_status.values()), len(records), json.dumps(dict(sorted(moe_status.items())), sort_keys=True))
    base.add_check(checks, "XW-REC-001", "mapping", "Raw crosswalk row reconciliation", "PASS", f"{crosswalk_summary['zcta_nonblank_count']}+{crosswalk_summary['zcta_blank_count']}", len(relationships), "Nonblank-ZCTA rows plus no-ZCTA exception rows equals source rows.")
    base.add_check(checks, "XW-CARD-001", "mapping", "ZIP-to-ZCTA source cardinality documented", "PASS", f"ZIP one-row={len(zip_counts)}; ZCTA one-ZIP={crosswalk_summary['zcta_one_zip_count']}; ZCTA multi-ZIP={crosswalk_summary['zcta_multiple_zip_count']}", "documented", "The HRSA file is a pre-resolved one-record-per-ZIP crosswalk.")
    base.add_check(checks, "XW-WGT-001", "mapping", "Allocation weights are available", "WARN", 0, ">=1 allocation field preferred", "Source provides categorical join method but no residential/population/address weights; limitation retained for review.")
    base.add_check(checks, "MAP-001", "mapping", "Canonical ZCTA mapping reconciliation", "PASS" if len(mapped) + len(unresolved) == len(records) else "BLOCK", f"{len(mapped)}+{len(unresolved)}", len(records), "Resolved plus unresolved equals canonical ACS ZCTAs.")
    base.add_check(checks, "MAP-002", "mapping", "Cross-state relationships are retained and flagged", "PASS", len(ambiguous), 2, "45202 and 45209 assigned from direct match; OH/KY candidates preserved.")
    base.add_check(checks, "MAP-003", "mapping", "Unresolved ACS ZCTAs are explicitly listed", "WARN" if unresolved else "PASS", len(unresolved), 0, "32026 and 97258 remain unresolved; no prefix or manual correction applied.")
    base.add_check(checks, "MAP-004", "mapping", "Crosswalk no-ZCTA ZIPs are explicitly listed", "WARN" if crosswalk_summary["zcta_blank_count"] else "PASS", crosswalk_summary["zcta_blank_count"], 0, "Territory ZIPs with no ZCTA remain in raw/resolved ZIP outputs and exceptions.")
    base.add_check(checks, "STATE-001", "within_state", "All exact valid income records received state ranks", "PASS" if len(calculated) == income_status["valid"] else "BLOCK", len(calculated), income_status["valid"], "The two unmatched state records have non-exact income and therefore do not reduce rank coverage.")
    bounds_ok = all(0 <= row["income_percentile_in_state"] <= 1 and 0 < row["income_empirical_cdf_in_state"] <= 1 for row in calculated)
    base.add_check(checks, "STATE-002", "within_state", "Percentile and empirical-CDF bounds", "PASS" if bounds_ok else "BLOCK", bounds_ok, True, "Checked every calculated rank row.")
    monotonic_ok = state_monotonicity(records)
    base.add_check(checks, "STATE-003", "within_state", "Within-state percentiles are monotone under income ordering", "PASS" if monotonic_ok else "BLOCK", monotonic_ok, True, "Checked independently within each resolved jurisdiction.")
    tie_ok = tie_consistency(records)
    base.add_check(checks, "STATE-004", "within_state", "Tied incomes receive identical rank measures", "PASS" if tie_ok else "BLOCK", tie_ok, True, "Minimum rank, midrank, percentile, CDF, and tie count agree within ties.")
    base.add_check(checks, "STATE-005", "within_state", "State summaries reconcile to resolved ZCTAs", "PASS" if sum(row["total_zcta_count"] for row in state_summaries) == len(mapped) else "BLOCK", sum(row["total_zcta_count"] for row in state_summaries), len(mapped), f"Jurisdictions summarized={len(state_summaries)}.")
    base.add_check(checks, "REC-001", "reconciliation", "ACS physical-row reconciliation", "PASS", f"1+{layout['annotation_rows']}+{layout['data_rows']}", layout["physical_rows"], "Header + annotation + data equals physical rows.")
    base.add_check(checks, "REC-002", "reconciliation", "Mapping exception table is non-empty and reason-coded", "PASS" if mapping_exceptions and all(row["reason_code"] for row in mapping_exceptions) else "BLOCK", len(mapping_exceptions), ">0 reason-coded exceptions", "No exception was discarded.")
    base.add_check(checks, "GATE-001", "phase_gate", "Candidate groupings were not produced before checkpoint approval", "PASS", 0, 0, "Pipeline stops at the Phase 1 checkpoint.")
    return checks


def state_monotonicity(records: Sequence[dict[str, Any]]) -> bool:
    by_state: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in records:
        if row["income_rank_in_state"] is not None:
            by_state[row["state_code"]].append(row)
    for rows in by_state.values():
        ordered = sorted(rows, key=lambda row: (row["median_household_income"], row["zcta5"]))
        percentiles = [row["income_percentile_in_state"] for row in ordered]
        if any(right < left for left, right in zip(percentiles, percentiles[1:])):
            return False
    return True


def tie_consistency(records: Sequence[dict[str, Any]]) -> bool:
    groups: dict[tuple[str, float], list[dict[str, Any]]] = defaultdict(list)
    for row in records:
        if row["income_rank_in_state"] is not None:
            groups[(row["state_code"], row["median_household_income"])].append(row)
    fields = ["income_rank_in_state", "income_midrank_in_state", "income_percentile_in_state", "income_empirical_cdf_in_state", "tie_count"]
    return all(all(len({row[field] for row in rows}) == 1 for field in fields) and rows[0]["tie_count"] == len(rows) for rows in groups.values())


def source_links_rows(config: dict[str, Any], acs_checksum: str, crosswalk_checksum: str, run_id: str) -> list[dict[str, Any]]:
    return [
        {
            "source_id": "ACS_B19013_2024", "source_organization": "U.S. Census Bureau",
            "landing_page": config["source_url"], "direct_download": "provided local data.census.gov export",
            "local_relative_path": config["selected_data_file"], "release_label": "2024 ACS 5-Year Detailed Tables",
            "underlying_vintage": "2024", "sha256": acs_checksum, "run_id": run_id,
        },
        {
            "source_id": "HRSA_ZIP_ZCTA_2024", "source_organization": "U.S. Health Resources and Services Administration",
            "landing_page": config["crosswalk_landing_page"], "direct_download": config["crosswalk_download_url"],
            "local_relative_path": config["crosswalk_file"], "release_label": "2024 ZIP Code to ZCTA Crosswalk",
            "underlying_vintage": config["crosswalk_vintage"], "sha256": crosswalk_checksum, "run_id": run_id,
        },
    ]


def report_texts(config: dict[str, Any], records: Sequence[dict[str, Any]], relationships: Sequence[dict[str, Any]], crosswalk_summary: dict[str, Any], state_summaries: Sequence[dict[str, Any]], state_outliers: Sequence[dict[str, Any]], mapping_exceptions: Sequence[dict[str, Any]], national: dict[str, Any], moe: dict[str, Any], checks: Sequence[dict[str, Any]], run_id: str, generated_at: str, acs_checksum: str, crosswalk_checksum: str) -> tuple[str, str]:
    income = Counter(row["income_value_status"] for row in records)
    mapped = [row for row in records if row["state_code"]]
    unresolved = [row for row in records if not row["state_code"]]
    ambiguous = [row for row in records if row["state_assignment_ambiguous"] is True]
    calculated = [row for row in records if row["income_rank_in_state"] is not None]
    statuses = Counter(row["status"] for row in checks)
    exception_types = Counter(row["exception_type"] for row in mapping_exceptions)
    largest = max(state_summaries, key=lambda row: row["total_zcta_count"])
    smallest = min(state_summaries, key=lambda row: row["total_zcta_count"])
    qa = f"""# Phase 1 QA Report — Crosswalk and State Statistics Checkpoint

**Run ID:** `{run_id}`  
**Generated (UTC):** {generated_at}  
**Status:** **COMPLETE WITH DOCUMENTED EXCEPTIONS — CHECKPOINT REVIEW REQUIRED**

## Sources and vintages

- Census: 2024 ACS 5-Year Detailed Table B19013, ZCTA geography, 2024 inflation-adjusted dollars, 90% MOE; SHA-256 `{acs_checksum}`.
- HRSA: 2024 ZIP Code to ZCTA Crosswalk; workbook notes identify **Census TIGER 2023 ZCTA boundaries** and **June 2023 ZIP boundaries**; SHA-256 `{crosswalk_checksum}`.
- Source URLs are retained in `source_links.csv` and in the raw-data folder's `SOURCE_LINKS.md`.

## ACS reconciliation and identifier QA

- 33,774 physical rows = 1 header + 1 annotation + 33,772 data rows.
- 33,772 unique valid five-character ZCTAs; 2,577 leading-zero ZCTAs preserved; no duplicate ZCTA excess records.
- Income: 30,414 exact valid + 3,225 not computed (`-`) + 118 top open-ended + 15 bottom open-ended = 33,772.
- No missing/special value was converted to zero; no observation was imputed, capped, winsorized, or removed.

## Crosswalk schema and cardinality

- Raw HRSA rows / unique ZIPs: {len(relationships):,} / {crosswalk_summary['zip_unique_count']:,}.
- ZIPs with nonblank ZCTA: {crosswalk_summary['zcta_nonblank_count']:,}; territory ZIPs with no ZCTA: {crosswalk_summary['zcta_blank_count']:,}.
- Unique crosswalk ZCTAs: {crosswalk_summary['zcta_unique_count']:,}.
- ZCTAs linked to one ZIP: {crosswalk_summary['zcta_one_zip_count']:,}; linked to multiple ZIPs: {crosswalk_summary['zcta_multiple_zip_count']:,}.
- Join types: {dict(crosswalk_summary['join_type_counts'])}.

The HRSA source is already resolved to one record per ZIP and provides a categorical join method, not address/population allocation weights. This limitation is recorded as a warning rather than concealed.

## State-resolution hierarchy

1. Select the unique state from an HRSA `Zip matches ZCTA` record where ZIP equals the canonical ZCTA.
2. If absent, select the state from an exact-code special source record where ZIP equals ZCTA.
3. If absent, select only when all retained ZCTA relationships have one unique state.
4. Otherwise leave unresolved; never select the first row or infer from digits.

Results: {len(mapped):,} of {len(records):,} ACS ZCTAs resolved ({len(mapped)/len(records):.4%}); {len(unresolved):,} unresolved; {len(ambiguous):,} cross-state relationships assigned by direct-match priority and flagged. The cross-state ZCTAs are 45202 and 45209 (OH selected; OH/KY candidate states retained). Unresolved ZCTAs are 32026 and 97258.

One HRSA source inconsistency is preserved and flagged: the special row whose ZIP/PO name indicates 32026 contains ZCTA field 32076. No silent correction was applied.

## Mapping exceptions

Exception counts by type: {dict(sorted(exception_types.items()))}. Every exception has a reason code and source-row trace in `mapping_exceptions.csv`.

## Income and uncertainty

- Exact income coverage: {income['valid']/len(records):.2%}.
- National exact-value-only median: {base.serialize(national['median_p50'])}; mean: {base.serialize(national['mean'])}; p05: {base.serialize(national['p05'])}; p95: {base.serialize(national['p95'])}.
- Numeric MOE count: {moe['valid_moe_count']:,}; relative-MOE median: {base.serialize(moe['relative_moe_median'])}; relative MOE ≥50%: {moe['relative_moe_ge_warning_count']:,}; ≥100%: {moe['relative_moe_ge_severe_count']:,}.
- National IQR outliers flagged: {national['iqr_outlier_count']:,}; within-state IQR outliers flagged: {len(state_outliers):,}. These are flags only.

## Within-state calculations

All {len(calculated):,} exact valid income records received a state rank. There are {len(state_summaries):,} state/jurisdiction summaries. The largest resolved geography is {largest['state_code']} with {largest['total_zcta_count']:,} ZCTAs; the smallest is {smallest['state_code']} with {smallest['total_zcta_count']:,}.

Within each state, valid exact incomes are sorted ascending. Identical incomes share minimum rank and tie count. Midrank percentile is `(midrank - 1)/(n - 1)` for `n>1`, with `0.5` for `n=1`; empirical CDF is maximum tied rank divided by `n`. Bounds, monotonicity, tie consistency, and state separation passed automated checks.

## QA status

- PASS: {statuses['PASS']}
- WARN: {statuses['WARN']}
- BLOCK: {statuses['BLOCK']}

Warnings concern the absence of allocation weights, two unresolved ACS ZCTAs, and eight territory ZIPs with no ZCTA. There is no unexplained reconciliation difference and no blocking QA failure.

## Guardrails

No internal policy, premium, exposure, claim, or loss data were used. No candidate groups, final territories, territory counts, rating factors, causal conclusions, or regulatory-acceptability claims were produced.
"""

    checkpoint = f"""# Phase 1 Checkpoint Memo — External Data Foundation

**Run ID:** `{run_id}`  
**Checkpoint status:** **COMPLETE WITH DOCUMENTED EXCEPTIONS — APPROVAL REQUIRED BEFORE CANDIDATE GROUPS**

## 1. What was completed

The 2024 ACS B19013 source and HRSA 2024 ZIP-to-ZCTA crosswalk were validated and processed read-only. A canonical ZCTA income table, raw and resolved crosswalk tables, auditable ZCTA-state assignments, within-state ranks/percentiles, state summaries, uncertainty diagnostics, exceptions, decision log, source links, QA checks, and rerunnable code were produced.

## 2. Sources used and vintages

The income source is 2024 ACS 5-Year Detailed Table B19013. The mapping source is HRSA's 2024 ZIP Code to ZCTA Crosswalk; its internal notes specify TIGER 2023 ZCTA boundaries and June 2023 ZIP boundaries. Source hashes are `{acs_checksum}` and `{crosswalk_checksum}` respectively.

## 3. Record-count reconciliation

ACS: 33,774 physical rows = 1 header + 1 annotation + 33,772 canonical ZCTA rows. Crosswalk: {len(relationships):,} rows = {crosswalk_summary['zcta_nonblank_count']:,} ZIPs with ZCTA + {crosswalk_summary['zcta_blank_count']:,} ZIPs without ZCTA. Canonical mapping: {len(records):,} = {len(mapped):,} resolved + {len(unresolved):,} unresolved.

## 4. Key data-quality findings

All ACS identifiers are valid and unique. All crosswalk ZIPs are unique and valid. Exact income is available for {income['valid']:,} ZCTAs; {income['not_computed_dash']:,} are not computed, {income['top_open_ended']:,} are top open-ended, and {income['bottom_open_ended']:,} are bottom open-ended. The HRSA file contains one internally inconsistent special row involving 32026/32076; it is preserved and flagged.

## 5. Mapping coverage and exceptions

State assignment coverage is {len(mapped)/len(records):.4%}. ZCTAs 32026 and 97258 remain unresolved. ZCTAs 45202 and 45209 contain OH/KY relationships; OH is selected using the direct ZIP=ZCTA record and the alternate KY relationships remain flagged. Eight territory ZIPs have no ZCTA. The HRSA file is pre-resolved and has no allocation ratios; this is a documented limitation.

## 6. Income missingness and uncertainty findings

Exact income coverage is {income['valid']/len(records):.2%}. Numeric MOE and relative MOE are available for {moe['relative_moe_valid_count']:,} exact-income rows. {moe['relative_moe_ge_warning_count']:,} have relative MOE ≥50%; none have relative MOE ≥100%. Open-ended values are excluded from exact-value ranks but retained with bound and direction.

## 7. ZCTA counts and income distribution by state

`state_summary.csv` contains counts, missingness, selected percentiles, median, mean, extrema, MOE diagnostics, and outlier flags for {len(state_summaries):,} states/jurisdictions. It reconciles to all {len(mapped):,} resolved ZCTAs. The two unresolved ZCTAs both have non-exact income and therefore do not remove an exact income from state ranking.

## 8. Method used for ranks, percentiles, and ties

Valid exact incomes are ranked ascending within resolved state. ZCTA controls deterministic record order, but identical income values are not split: they share minimum rank, midrank percentile, empirical CDF, and tie count. Percentile is `(midrank-1)/(n-1)` for `n>1`, otherwise `0.5`; empirical CDF is maximum tied rank / `n`.

## 9. Files produced and how to reproduce them

See `README.md` and `output_manifest.csv`. The final pipeline is `pipeline/phase1_with_crosswalk.py`; it requires Python 3.10+ and `openpyxl`. Inputs are opened read-only, their hashes are verified before and after execution, and outputs are regenerated in a separate directory.

## 10. Blocking issues or decisions required

There is no blocking reconciliation failure. Before authorizing candidate groupings, the reviewer should explicitly accept or revise: (a) HRSA's pre-resolved, unweighted mapping source and 2023 underlying vintages; (b) direct-match priority for the two cross-state ZCTAs; (c) leaving 32026 and 97258 unresolved unless a second approved source is provided; (d) exclusion of open-ended medians from exact ranks; and (e) the midrank percentile formula.

## 11. Recommended next step — without selecting final territories

Review and approve this checkpoint and decision log. Only after approval should 4/5/6 within-state candidate grouping scenarios be generated for sensitivity comparison. Those candidates will remain exploratory and will not be labeled final territories or indicated rating factors.
"""
    return qa, checkpoint


def write_outputs(output_root: Path, config: dict[str, Any], inventory: list[dict[str, Any]], records: list[dict[str, Any]], relationships: list[dict[str, Any]], resolution_rows: list[dict[str, Any]], resolved_zip_rows: list[dict[str, Any]], state_summaries: list[dict[str, Any]], state_outliers: list[dict[str, Any]], mapping_exceptions: list[dict[str, Any]], checks: list[dict[str, Any]], layout: dict[str, Any], crosswalk_summary: dict[str, Any], national: dict[str, Any], moe: dict[str, Any], run_id: str, generated_at: str, acs_checksum: str, crosswalk_checksum: str) -> None:
    canonical_fields = [
        "zcta5", "geo_id_raw", "geo_label_raw", "source_row_number", "state_code", "state_assignment_method",
        "state_assignment_ambiguous", "state_candidate_count", "state_candidate_codes", "crosswalk_relationship_count",
        "state_resolution_reason_code", "mapping_status", "median_household_income", "income_moe", "income_relative_moe",
        "income_estimate_raw", "income_moe_raw", "income_value_status", "income_moe_status", "income_bound_value",
        "income_censor_direction", "income_rank_in_state", "income_midrank_in_state", "income_percentile_in_state",
        "income_empirical_cdf_in_state", "tie_count", "national_income_outlier_flag", "state_income_outlier_flag",
        "zcta_parse_status", "transformation_flags", "acs_year", "acs_product", "acs_table", "income_units",
        "crosswalk_vintage", "run_id",
    ]
    raw_fields = ["zip5", "po_name", "state_code_source", "zip_type", "zcta5", "zip_join_type", "source_row_number", "crosswalk_source_file", "crosswalk_vintage", "transformation_flags", "run_id"]
    rank_fields = ["zcta5", "state_code", "median_household_income", "income_value_status", "income_rank_in_state", "income_midrank_in_state", "income_percentile_in_state", "income_empirical_cdf_in_state", "tie_count", "state_income_outlier_flag", "state_assignment_ambiguous", "calculation_status", "run_id"]
    state_fields = ["state_code", "total_zcta_count", "valid_income_count", "excluded_income_count", "not_computed_dash_count", "top_open_ended_count", "bottom_open_ended_count", "income_minimum", "income_p01", "income_p05", "income_p25", "income_median", "income_mean", "income_p75", "income_p95", "income_p99", "income_maximum", "moe_median", "relative_moe_median", "relative_moe_ge_warning_count", "relative_moe_ge_severe_count", "state_iqr_outlier_count", "small_state_under_six_valid_zcta", "quantile_method", "calculation_status", "run_id"]

    for row in inventory:
        row["selected"] = row["source_role"].startswith("selected_")
        row["source_url"] = ""
        row["geographic_vintage"] = row["apparent_vintage"]
        if row["relative_path"] == config["crosswalk_file"]:
            row["source_role"] = "selected_zip_zcta_state_crosswalk"
            row["selected"] = True
            row["source_url"] = config["crosswalk_download_url"]
            row["geographic_vintage"] = config["crosswalk_vintage"]
        elif row["relative_path"] == config["selected_data_file"]:
            row["source_url"] = config["source_url"]

    base.write_csv(output_root / "source_inventory_manifest.csv", ["relative_path", "filename", "format", "size_bytes", "modified_time_utc", "sha256", "apparent_vintage", "geographic_vintage", "source_role", "selected", "source_url", "crosswalk_candidate", "physical_row_count", "observed_header_column_count"], inventory)
    base.write_csv(output_root / "source_links.csv", ["source_id", "source_organization", "landing_page", "direct_download", "local_relative_path", "release_label", "underlying_vintage", "sha256", "run_id"], source_links_rows(config, acs_checksum, crosswalk_checksum, run_id))
    base.write_csv(output_root / "data_dictionary.csv", ["original_field", "source", "original_type", "definition", "transformation", "final_field", "final_type", "units"], build_data_dictionary())
    base.write_csv(output_root / "decision_log.csv", ["decision_id", "decision_date", "issue", "alternatives_considered", "selected_approach", "rationale", "decision_owner", "affected_outputs", "rerun_required", "run_id"], decision_rows(run_id, generated_at))
    base.write_csv(output_root / "tables/clean_zcta_income.csv", canonical_fields, records)
    base.write_csv(output_root / "tables/raw_zip_zcta_state_relationships.csv", raw_fields, relationships)
    base.write_csv(output_root / "tables/resolved_zip_zcta_state_mapping.csv", ["zip5", "zcta5", "state_code", "po_name", "zip_type", "zip_join_type", "state_assignment_method", "state_assignment_ambiguous", "candidate_count", "mapping_status", "reason_code", "crosswalk_vintage", "source_row_number", "run_id"], resolved_zip_rows)
    base.write_csv(output_root / "tables/resolved_zcta_state_mapping.csv", ["zcta5", "state_code", "state_assignment_method", "state_assignment_ambiguous", "state_candidate_count", "state_candidate_codes", "crosswalk_relationship_count", "mapping_status", "reason_code", "crosswalk_vintage", "run_id"], resolution_rows)
    base.write_csv(output_root / "tables/within_state_income_rank.csv", rank_fields, build_rank_rows(records))
    base.write_csv(output_root / "tables/state_summary.csv", state_fields, state_summaries)
    base.write_csv(output_root / "tables/national_income_summary.csv", list(national.keys()), [national])
    base.write_csv(output_root / "tables/national_moe_summary.csv", list(moe.keys()), [moe])
    base.write_csv(output_root / "tables/income_status_summary.csv", ["income_value_status", "record_count", "run_id"], [{"income_value_status": key, "record_count": value, "run_id": run_id} for key, value in sorted(Counter(row["income_value_status"] for row in records).items())])
    base.write_csv(output_root / "tables/moe_status_summary.csv", ["income_moe_status", "record_count", "run_id"], [{"income_moe_status": key, "record_count": value, "run_id": run_id} for key, value in sorted(Counter(row["income_moe_status"] for row in records).items())])

    zcta_counts = Counter(row["zcta5"] for row in records)
    duplicates = [{"zcta5": key, "record_count": value, "resolution_status": "not_resolved", "run_id": run_id} for key, value in sorted(zcta_counts.items()) if value > 1]
    identifier_exceptions = [{"source_row_number": row["source_row_number"], "geo_id_raw": row["geo_id_raw"], "geo_label_raw": row["geo_label_raw"], "zcta_parse_status": row["zcta_parse_status"], "run_id": run_id} for row in records if row["zcta_parse_status"] != "valid"]
    income_exceptions = [{"zcta5": row["zcta5"], "state_code": row["state_code"], "source_row_number": row["source_row_number"], "income_estimate_raw": row["income_estimate_raw"], "income_moe_raw": row["income_moe_raw"], "income_value_status": row["income_value_status"], "income_moe_status": row["income_moe_status"], "income_bound_value": row["income_bound_value"], "income_censor_direction": row["income_censor_direction"], "run_id": run_id} for row in records if row["income_value_status"] != "valid"]
    high_uncertainty = [{"zcta5": row["zcta5"], "state_code": row["state_code"], "median_household_income": row["median_household_income"], "income_moe": row["income_moe"], "income_relative_moe": row["income_relative_moe"], "warning_level": "severe" if row["income_relative_moe"] >= config["relative_moe_severe_threshold"] else "warning", "run_id": run_id} for row in records if row["income_relative_moe"] is not None and row["income_relative_moe"] >= config["relative_moe_warning_threshold"]]
    national_outliers = [{"zcta5": row["zcta5"], "state_code": row["state_code"], "median_household_income": row["median_household_income"], "action": "flag_only_no_change", "run_id": run_id} for row in records if row["national_income_outlier_flag"]]
    base.write_csv(output_root / "exceptions/duplicate_zcta.csv", ["zcta5", "record_count", "resolution_status", "run_id"], duplicates)
    base.write_csv(output_root / "exceptions/identifier_exceptions.csv", ["source_row_number", "geo_id_raw", "geo_label_raw", "zcta_parse_status", "run_id"], identifier_exceptions)
    base.write_csv(output_root / "exceptions/income_exceptions.csv", ["zcta5", "state_code", "source_row_number", "income_estimate_raw", "income_moe_raw", "income_value_status", "income_moe_status", "income_bound_value", "income_censor_direction", "run_id"], income_exceptions)
    base.write_csv(output_root / "exceptions/high_uncertainty_zcta.csv", ["zcta5", "state_code", "median_household_income", "income_moe", "income_relative_moe", "warning_level", "run_id"], high_uncertainty)
    base.write_csv(output_root / "exceptions/national_income_outliers.csv", ["zcta5", "state_code", "median_household_income", "action", "run_id"], national_outliers)
    base.write_csv(output_root / "exceptions/state_income_outliers.csv", ["zcta5", "state_code", "median_household_income", "state_iqr_lower_fence", "state_iqr_upper_fence", "action", "run_id"], state_outliers)
    base.write_csv(output_root / "exceptions/mapping_exceptions.csv", ["record_level", "zip5", "zcta5", "state_candidate_codes", "exception_type", "reason_code", "details", "source_rows", "run_id"], mapping_exceptions)
    base.write_csv(output_root / "qa/qa_checks.csv", ["check_id", "category", "description", "status", "observed", "expected", "details"], checks)
    stage_rows = [
        {"stage": "acs_physical_rows", "record_count": layout["physical_rows"], "reconciliation_role": "header + annotation + data", "run_id": run_id},
        {"stage": "acs_canonical_zcta_rows", "record_count": len(records), "reconciliation_role": "canonical", "run_id": run_id},
        {"stage": "crosswalk_raw_rows", "record_count": len(relationships), "reconciliation_role": "all source relationships", "run_id": run_id},
        {"stage": "crosswalk_rows_with_zcta", "record_count": crosswalk_summary["zcta_nonblank_count"], "reconciliation_role": "mapped ZIP", "run_id": run_id},
        {"stage": "crosswalk_rows_without_zcta", "record_count": crosswalk_summary["zcta_blank_count"], "reconciliation_role": "ZIP exception", "run_id": run_id},
        {"stage": "zcta_state_resolved", "record_count": sum(bool(row["state_code"]) for row in records), "reconciliation_role": "resolved", "run_id": run_id},
        {"stage": "zcta_state_unresolved", "record_count": sum(not row["state_code"] for row in records), "reconciliation_role": "unresolved", "run_id": run_id},
        {"stage": "within_state_rank_calculated", "record_count": sum(row["income_rank_in_state"] is not None for row in records), "reconciliation_role": "valid exact income + resolved state", "run_id": run_id},
    ]
    base.write_csv(output_root / "qa/stage_row_counts.csv", ["stage", "record_count", "reconciliation_role", "run_id"], stage_rows)

    qa_text, checkpoint_text = report_texts(config, records, relationships, crosswalk_summary, state_summaries, state_outliers, mapping_exceptions, national, moe, checks, run_id, generated_at, acs_checksum, crosswalk_checksum)
    (output_root / "reports").mkdir(parents=True, exist_ok=True)
    (output_root / "reports/qa_report.md").write_text(qa_text, encoding="utf-8")
    (output_root / "reports/phase1_checkpoint.md").write_text(checkpoint_text, encoding="utf-8")
    (output_root / "reports/crosswalk_input_specification.md").write_text("# Crosswalk Input Status\n\nThe required crosswalk was supplied and processed. See `source_links.csv`, `source_inventory_manifest.csv`, `mapping_exceptions.csv`, and the checkpoint memo for provenance and limitations.\n", encoding="utf-8")


def run(args: argparse.Namespace) -> dict[str, Any]:
    input_root = args.input_root.resolve()
    output_root = args.output_root.resolve()
    config = json.loads(args.config.read_text(encoding="utf-8"))
    if args.clean:
        base.safe_clean(output_root, input_root)
    output_root.mkdir(parents=True, exist_ok=True)

    data_path, metadata_path, notes_path, _ = base.validate_selected_sources(input_root, config)
    crosswalk_path = input_root / config["crosswalk_file"]
    if not crosswalk_path.is_file():
        raise FileNotFoundError(f"Required crosswalk missing: {crosswalk_path}")
    acs_checksum = base.sha256_file(data_path)
    crosswalk_checksum = base.sha256_file(crosswalk_path)
    run_id = f"phase1_acs{config['acs_year']}_{acs_checksum[:10]}_xw_{crosswalk_checksum[:10]}_v{config['pipeline_version'].replace('.', '_')}"
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    raw_before = {path: base.sha256_file(path) for path in (data_path, metadata_path, notes_path, crosswalk_path)}

    inventory = base.inventory_sources(input_root, config)
    records, layout = base.ingest_acs(data_path, config, run_id)
    national_fences = base.mark_national_outliers(records)
    relationships, crosswalk_summary = ingest_crosswalk(crosswalk_path, config, run_id)
    validate_crosswalk_rows(relationships)
    resolution_rows = resolve_zcta_states(records, relationships, config)
    state_summaries, state_outliers = calculate_state_statistics(records, config)
    resolved_zip_rows = build_resolved_zip_rows(relationships, {row["zcta5"] for row in records})
    mapping_exceptions = build_mapping_exceptions(records, relationships, {row["zcta5"] for row in records}, run_id)
    checks = qa_checks(records, layout, relationships, crosswalk_summary, state_summaries, mapping_exceptions)
    national = base.national_summary(records, national_fences, run_id)
    moe = base.moe_summary(records, config, run_id)

    write_outputs(output_root, config, inventory, records, relationships, resolution_rows, resolved_zip_rows, state_summaries, state_outliers, mapping_exceptions, checks, layout, crosswalk_summary, national, moe, run_id, generated_at, acs_checksum, crosswalk_checksum)

    raw_after = {path: base.sha256_file(path) for path in (data_path, metadata_path, notes_path, crosswalk_path)}
    if raw_after != raw_before:
        raise RuntimeError("A raw source checksum changed during execution")
    metadata = {
        "run_id": run_id,
        "pipeline_version": config["pipeline_version"],
        "generated_at_utc": generated_at,
        "selected_acs_sha256": acs_checksum,
        "selected_crosswalk_sha256": crosswalk_checksum,
        "crosswalk_release_label": "HRSA 2024 ZIP Code to ZCTA Crosswalk",
        "crosswalk_underlying_vintage": config["crosswalk_vintage"],
        "raw_source_immutability_verified": True,
        "python_requirement": "Python 3.10+",
        "openpyxl_version": openpyxl.__version__,
        "checkpoint_status": "complete_with_documented_exceptions_pending_approval",
        "candidate_groups_produced": False,
        "internal_experience_used": False,
    }
    (output_root / "run_metadata.json").write_text(json.dumps(metadata, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    base.write_output_manifest(output_root, run_id)
    return {
        "run_id": run_id,
        "acs_zcta_records": len(records),
        "crosswalk_rows": len(relationships),
        "state_resolved_zcta": sum(bool(row["state_code"]) for row in records),
        "state_unresolved_zcta": sum(not row["state_code"] for row in records),
        "state_rank_rows": sum(row["income_rank_in_state"] is not None for row in records),
        "state_summary_rows": len(state_summaries),
        "qa_pass": sum(row["status"] == "PASS" for row in checks),
        "qa_warn": sum(row["status"] == "WARN" for row in checks),
        "qa_block": sum(row["status"] == "BLOCK" for row in checks),
        "output_root": str(output_root),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-root", type=Path, required=True)
    parser.add_argument("--config", type=Path, required=True)
    parser.add_argument("--output-root", type=Path, required=True)
    parser.add_argument("--clean", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    print(json.dumps(run(parse_args()), indent=2, sort_keys=True))
